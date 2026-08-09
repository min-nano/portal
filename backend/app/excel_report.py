"""傾斜測定報告書（Excel）の生成。

gas-addon-excel-report-formatter の Cloud Function（functions/main.py）から
移植した生成ロジック。書き込み位置（セルマッピング）は mapping.json に
切り出しており、フォーマットが微修正された場合は原則 mapping.json を
編集するだけで追従できる。
"""

import io
import json
import os
import unicodedata

from openpyxl import load_workbook

_MAPPING_PATH = os.path.join(os.path.dirname(__file__), "mapping.json")
_MAPPING = None

REPORT_FILE_NAME = "傾斜測定報告書.xlsx"


class ReportError(Exception):
    """入力起因の生成エラー。message は利用者に表示できる日本語文。"""

    def __init__(self, message: str, status: int = 400):
        super().__init__(message)
        self.status = status


def load_mapping() -> dict:
    # mapping.json は静的な設定ファイルなので、一度読み込んだらキャッシュする。
    # Cloud Run のインスタンスはリクエスト間で再利用されるため、毎回の
    # ディスク I/O を避けられる。
    global _MAPPING
    if _MAPPING is None:
        with open(_MAPPING_PATH, encoding="utf-8") as f:
            _MAPPING = json.load(f)
    return _MAPPING


def form_config() -> dict:
    """フロントエンドがフォームを組み立てるための設定を mapping.json から導出する。

    GAS 版ではフロントエンドの MEASUREMENT_GROUPS / VALIDATION 定数を
    mapping.json と手動で一致させる必要があったが、本ポータルでは API で
    配信することで mapping.json を単一の情報源にする。
    """
    mapping = load_mapping()
    groups: list[dict] = []
    for m in mapping["room_block"]["measurements"]:
        select = m.get("select", {})
        if not groups or groups[-1]["group"] != m["group"]:
            groups.append(
                {
                    "group": m["group"],
                    "select_label": select.get("label", ""),
                    "points": [],
                }
            )
        groups[-1]["points"].append(
            {
                "key": m["key"],
                "label": m["label"],
                "options": select.get("options", []),
            }
        )
    return {
        "measurement_groups": groups,
        "validation": mapping["validation"],
        "max_rooms": len(mapping["block_start_rows"]),
        "report_file_name": REPORT_FILE_NAME,
    }


def _to_cell_value(value):
    """フォームから来る値を Excel セル向けに整える。

    数値として解釈できる文字列は数値に変換する（換算計測値 AJ の数式
    =1000*AC/AG が正しく計算されるよう、測定値の差・距離などは数値で書き込む）。
    空文字・None は None（空セル）として扱う。
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text == "":
        return None
    # モバイルの日本語入力では全角の数字・記号（－ ． など）が混じりやすい。
    # NFKC 正規化で半角へ寄せてから数値判定することで、Excel 側で数値として
    # 認識されず数式が計算されない事故を防ぐ。
    text = unicodedata.normalize("NFKC", text)
    try:
        if text.lstrip("-").isdigit():
            return int(text)
        return float(text)
    except ValueError:
        return text


def _write_cell(ws, cell_ref, value):
    """シート保護ルールに従いセルへ書き込む。

    シート保護が有効かつセルが明示的にアンロック（locked=False）されていない場合は
    書き込みをスキップする。Excel のデフォルトではすべてのセルがロック状態のため、
    保護シートで編集可能にするには対象セルを事前に locked=False に設定しておく。
    シート保護が無効なテンプレートでは常に書き込む。
    """
    cell = ws[cell_ref]
    if ws.protection.sheet and cell.protection.locked is not False:
        return
    cell.value = value


def _clear_data_cells(ws, mapping):
    """全ブロックの「記入欄」を空にする。

    雛形には記入例（(例)ＬＤＫ など）が入っていることがあり、利用者が記入
    しなかった欄に例の値が残ると誤った報告書になる。そこで mapping.json が
    記入欄として定義しているセル（物件名・階数・部屋名・各計測点の選択欄と
    数値欄）だけを一度クリアしてから書き込む。印字済みのラベル・区切りの
    「/」・分母 1000・換算計測値の数式などは mapping に含まれないため消さない。
    シート保護が有効なセルはスキップする（_write_cell 参照）。
    """
    block = mapping["room_block"]
    value_fields = block["value_fields"]

    # property_name_cell を持たない将来フォーマットにも備え、定義時のみ消す。
    prop_cell = mapping.get("property_name_cell")
    if prop_cell:
        _write_cell(ws, prop_cell, None)
    for start_row in mapping["block_start_rows"]:
        _write_cell(ws, f"{block['floor_col']}{start_row}", None)
        _write_cell(ws, f"{block['room_name_col']}{start_row}", None)
        for m in block["measurements"]:
            row = start_row + m["row_offset"]
            select = m.get("select")
            if select:
                _write_cell(ws, f"{select['col']}{row}", None)
            for col in value_fields.values():
                _write_cell(ws, f"{col}{row}", None)


def _write_room(ws, mapping, start_row, room):
    """1 部屋ぶんのデータを、先頭行 start_row のブロックへ書き込む。"""
    # 不正な形（None や辞書以外）の要素は無視し、AttributeError を防ぐ。
    if not isinstance(room, dict):
        return

    block = mapping["room_block"]

    _write_cell(ws, f"{block['floor_col']}{start_row}", _to_cell_value(room.get("floor")))
    _write_cell(
        ws, f"{block['room_name_col']}{start_row}", _to_cell_value(room.get("room_name"))
    )

    value_fields = block["value_fields"]
    # measurements が辞書以外（リストや文字列など）でも落ちないようにする。
    measurements = room.get("measurements")
    if not isinstance(measurements, dict):
        measurements = {}
    for m in block["measurements"]:
        data = measurements.get(m["key"])
        if not isinstance(data, dict):
            continue
        row = start_row + m["row_offset"]
        # 選択欄（傾斜方向／測定した壁・柱）。雛形のプルダウン候補をそのまま
        # 書き込むため、数値化や正規化はせず文字列のまま入れる。
        select = m.get("select")
        if select and data.get("select") not in (None, ""):
            _write_cell(ws, f"{select['col']}{row}", data["select"])
        # 数値欄（水平器計測値・測定値の差・距離）。
        for field_key, col in value_fields.items():
            if field_key in data:
                _write_cell(ws, f"{col}{row}", _to_cell_value(data.get(field_key)))


def generate_report(template_bytes: bytes, property_name, rooms) -> bytes:
    """雛形（xlsx バイト列）にフォームデータを書き込み、xlsx バイト列を返す。"""
    mapping = load_mapping()

    # rooms は未指定なら空リスト。リスト以外は 400 で弾く。
    if rooms is None:
        rooms = []
    elif not isinstance(rooms, list):
        raise ReportError("rooms must be a list")

    block_start_rows = mapping["block_start_rows"]
    if len(rooms) > len(block_start_rows):
        raise ReportError(
            f"部屋数が雛形の上限（{len(block_start_rows)} 部屋）を超えています。"
            "雛形にブロックを追加し、mapping.json の block_start_rows を更新してください。"
        )

    wb = load_workbook(filename=io.BytesIO(template_bytes))
    ws = wb[mapping["sheet_name"]]

    # 雛形の記入例・前回値が残らないよう、記入欄を一度クリアする。
    _clear_data_cells(ws, mapping)

    # 物件名（共通項目）。property_name_cell を持たない将来フォーマットでも
    # mapping から該当キーを省くだけで対応できるよう、定義時のみ書き込む。
    prop_cell = mapping.get("property_name_cell")
    if prop_cell and property_name is not None:
        _write_cell(ws, prop_cell, _to_cell_value(property_name))

    # 各部屋のデータを、対応するブロックへ順に書き込む。
    for i, room in enumerate(rooms):
        _write_room(ws, mapping, block_start_rows[i], room)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
