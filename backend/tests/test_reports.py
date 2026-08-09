"""レポート生成 API（POST /api/tools/excel-report-formatter/reports）のテスト。

旧リポジトリ（gas-addon-excel-report-formatter/functions/test_main.py）の
テストを新 API に移植したもの。雛形はリクエストで受け取る代わりに Drive
（テストでは FakeDrive）から取得し、応答は Base64 JSON ではなく xlsx の
バイナリを直接返す。書き込み位置の期待値は mapping.json から解決する。
"""

from tests.util import (
    BLOCK,
    MAPPING,
    SHEET,
    STARTS,
    decode_workbook,
    make_protected_template_bytes,
    make_template_bytes,
    select_cell,
    value_cell,
)

REPORTS_URL = "/api/tools/excel-report-formatter/reports"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _post(client, drive, **fields):
    drive.configure_template()
    return client.post(REPORTS_URL, json=fields)


# --- 正常系 -----------------------------------------------------------------

def test_post_valid_data_returns_xlsx_binary(client, drive):
    resp = _post(
        client,
        drive,
        property_name="サンプル物件",
        rooms=[{"floor": "2", "room_name": "LDK", "measurements": {}}],
    )

    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(XLSX_MIME)
    # ファイル名は UTF-8 の filename* で返す。
    assert "filename*=UTF-8''" in resp.headers["content-disposition"]
    # XLSX は zip アーカイブなので PK シグネチャで始まる。
    assert resp.content[:2] == b"PK"


def test_template_is_fetched_as_the_requesting_user(client, drive):
    from tests.conftest import TEST_EMAIL

    _post(client, drive, rooms=[])

    # 雛形は Clerk JWT で確認したメールアドレスの代理セッションで取得される。
    assert drive.delegated_emails == [TEST_EMAIL]
    assert drive.fetch_calls == [("folder-1", "雛形.xlsx")]


def test_post_valid_data_writes_values_into_template(client, drive):
    resp = _post(
        client,
        drive,
        property_name="サンプル物件",
        rooms=[{
            "floor": "2",
            "room_name": "LDK",
            "measurements": {
                # 床: 傾斜方向は S 列へ
                "floor_x": {"select": "←", "diff": "0", "distance": "2000"},
                "floor_y": {"select": "↑", "diff": "3", "distance": "1500"},
                # 壁: 測定した壁は P 列へ
                "wall_ud": {"select": "上壁", "diff": "2", "distance": "1800"},
                # 柱: 計測できなかった場合は ―
                "pillar_lr": {"select": "―"},
            },
        }],
    )
    ws = decode_workbook(resp.content)[SHEET]

    # 共通項目
    assert ws[MAPPING["property_name_cell"]].value == "サンプル物件"
    # 部屋 0 のヘッダー
    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value == 2
    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value == "LDK"
    # 床 floor_x: 選択は S 列、数値文字列は数値化されて換算数式が計算できる
    assert ws[select_cell(0, "floor_x")].value == "←"
    assert ws[value_cell(0, "floor_x", "diff")].value == 0
    assert ws[value_cell(0, "floor_x", "distance")].value == 2000
    assert ws[select_cell(0, "floor_y")].value == "↑"
    # 壁: 選択は P 列（S 列ではない）
    assert ws[select_cell(0, "wall_ud")].value == "上壁"
    assert ws[value_cell(0, "wall_ud", "diff")].value == 2
    # 柱: 「―」（計測不可）はそのまま P 列へ
    assert ws[select_cell(0, "pillar_lr")].value == "―"


def test_fullwidth_numbers_are_normalized_to_numeric(client, drive):
    # モバイル IME は全角の数字・記号を入れがち。数値へ正規化されないと
    # 換算計測値の数式 (=1000*AC/AG) が計算できない。
    resp = _post(client, drive, rooms=[{
        "floor": "２", "room_name": "LDK",
        "measurements": {"floor_x": {"diff": "－３", "distance": "１５００"}},
    }])
    ws = decode_workbook(resp.content)[SHEET]

    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value == 2
    assert ws[value_cell(0, "floor_x", "diff")].value == -3
    assert ws[value_cell(0, "floor_x", "distance")].value == 1500


def test_non_dict_room_entries_are_skipped(client, drive):
    # rooms 配列に null が混ざってもクラッシュしない。
    resp = _post(client, drive, rooms=[
        None,
        {"floor": "2", "room_name": "寝室", "measurements": {}},
    ])
    assert resp.status_code == 200
    ws = decode_workbook(resp.content)[SHEET]
    assert ws[f"{BLOCK['room_name_col']}{STARTS[1]}"].value == "寝室"


def test_non_dict_measurements_are_ignored(client, drive):
    # measurements の型が不正でもクラッシュしない（リグレッションガード）。
    resp = _post(client, drive, rooms=[
        {"floor": "1", "room_name": "和室", "measurements": ["bogus"]},
    ])
    assert resp.status_code == 200
    ws = decode_workbook(resp.content)[SHEET]
    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value == "和室"


def test_multiple_rooms_map_to_successive_blocks(client, drive):
    resp = _post(client, drive, rooms=[
        {"floor": "1", "room_name": "玄関", "measurements": {}},
        {"floor": "2", "room_name": "寝室", "measurements": {}},
    ])
    ws = decode_workbook(resp.content)[SHEET]

    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value == "玄関"
    assert ws[f"{BLOCK['room_name_col']}{STARTS[1]}"].value == "寝室"


def test_empty_or_missing_fields_default_to_empty(client, drive):
    # 部屋も物件名も無し: 何も書き込まれないがファイルは生成される。
    resp = _post(client, drive, rooms=[])
    ws = decode_workbook(resp.content)[SHEET]

    assert ws[MAPPING["property_name_cell"]].value is None
    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value is None


def test_blank_measurement_fields_leave_cells_empty(client, drive):
    resp = _post(client, drive, rooms=[{
        "floor": "3", "room_name": "",
        "measurements": {"floor_x": {"select": "", "diff": "", "distance": ""}},
    }])
    ws = decode_workbook(resp.content)[SHEET]

    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value == 3
    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value is None
    assert ws[select_cell(0, "floor_x")].value is None
    assert ws[value_cell(0, "floor_x", "diff")].value is None


# --- シート保護 -------------------------------------------------------------

def test_locked_cells_in_protected_sheet_are_not_written(client, drive):
    # シート保護が有効でロックされたセル（入力例ブロック）へは書き込まない。
    drive.template_bytes = make_protected_template_bytes(unlocked_starts=STARTS[1:])
    resp = _post(client, drive, rooms=[
        {"floor": "2", "room_name": "LDK", "measurements": {}},
    ])
    assert resp.status_code == 200
    ws = decode_workbook(resp.content)[SHEET]
    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value is None
    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value is None


def test_unlocked_cells_in_protected_sheet_are_written(client, drive):
    drive.template_bytes = make_protected_template_bytes(unlocked_starts=STARTS[1:])
    resp = _post(client, drive, property_name="テスト物件", rooms=[
        None,  # 最初のブロックはロック済みのためスキップ
        {"floor": "3", "room_name": "寝室", "measurements": {
            "floor_x": {"select": "←", "diff": "2", "distance": "1500"},
        }},
    ])
    assert resp.status_code == 200
    ws = decode_workbook(resp.content)[SHEET]
    assert ws[MAPPING["property_name_cell"]].value == "テスト物件"
    assert ws[f"{BLOCK['floor_col']}{STARTS[1]}"].value == 3
    assert ws[f"{BLOCK['room_name_col']}{STARTS[1]}"].value == "寝室"
    assert ws[select_cell(1, "floor_x")].value == "←"
    assert ws[value_cell(1, "floor_x", "diff")].value == 2


def test_first_room_lands_in_a_writable_block(client, drive):
    # 回帰テスト: 実雛形では入力例ブロックは block_start_rows の外にあり、
    # 全ブロックが書き込み可能。1 部屋目が黙って失われないことを保証する。
    drive.template_bytes = make_protected_template_bytes(unlocked_starts=STARTS)
    resp = _post(client, drive, rooms=[
        {"floor": "1", "room_name": "和室", "measurements": {
            "floor_x": {"select": "←", "diff": "3", "distance": "1500"},
        }},
        {"floor": "2", "room_name": "洋室1", "measurements": {}},
    ])
    assert resp.status_code == 200
    ws = decode_workbook(resp.content)[SHEET]
    assert ws[f"{BLOCK['room_name_col']}{STARTS[0]}"].value == "和室"
    assert ws[select_cell(0, "floor_x")].value == "←"
    assert ws[value_cell(0, "floor_x", "diff")].value == 3
    assert ws[f"{BLOCK['room_name_col']}{STARTS[1]}"].value == "洋室1"


def test_locked_cells_in_protected_sheet_preserve_existing_content(client, drive):
    # ロックされたセルはクリアフェーズでも既存値（入力例）を保持する。
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Protection

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.protection.sheet = True
    ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value = "（例）"
    ws[MAPPING["property_name_cell"]].protection = Protection(locked=False)
    buf = io.BytesIO()
    wb.save(buf)
    drive.template_bytes = buf.getvalue()

    resp = _post(client, drive, property_name="サンプル物件", rooms=[])
    assert resp.status_code == 200
    ws = decode_workbook(resp.content)[SHEET]
    assert ws[f"{BLOCK['floor_col']}{STARTS[0]}"].value == "（例）"
    assert ws[MAPPING["property_name_cell"]].value == "サンプル物件"


# --- 入力エラー -------------------------------------------------------------

def test_rooms_not_a_list_returns_400(client, drive):
    resp = _post(client, drive, rooms=5)

    assert resp.status_code == 400
    assert resp.json() == {"error": "rooms must be a list"}


def test_too_many_rooms_returns_400(client, drive):
    rooms = [{"floor": "1", "room_name": f"部屋{i}", "measurements": {}}
             for i in range(len(STARTS) + 1)]
    resp = _post(client, drive, rooms=rooms)

    assert resp.status_code == 400
    assert "上限" in resp.json()["error"]


def test_post_without_body_returns_400(client, drive):
    drive.configure_template()
    resp = client.post(REPORTS_URL, content=b"", headers={"Content-Type": "application/json"})

    assert resp.status_code == 400
    assert resp.json() == {"error": "No data provided"}


def test_post_empty_json_object_returns_400(client, drive):
    drive.configure_template()
    resp = client.post(REPORTS_URL, json={})

    assert resp.status_code == 400
    assert resp.json() == {"error": "No data provided"}


def test_non_json_body_returns_400(client, drive):
    drive.configure_template()
    resp = client.post(REPORTS_URL, content=b"not-json", headers={"Content-Type": "text/plain"})

    assert resp.status_code == 400
    assert resp.json() == {"error": "No data provided"}


# --- 雛形まわりのエラー -----------------------------------------------------

def test_unconfigured_template_returns_409(client, drive):
    # 雛形が未設定のまま出力しようとした場合は設定を促すエラーを返す。
    resp = client.post(REPORTS_URL, json={"rooms": []})

    assert resp.status_code == 409
    assert "雛形が未設定" in resp.json()["error"]


def test_missing_template_file_returns_404(client, drive):
    # 設定はあるが Drive 上にファイルが見つからない（差し替え漏れなど）。
    drive.configure_template()
    drive.template_bytes = None
    resp = client.post(REPORTS_URL, json={"rooms": []})

    assert resp.status_code == 404
    assert "見つかりませんでした" in resp.json()["error"]


def test_template_without_expected_sheet_returns_500(client, drive):
    # 傾斜測定シートを持たないワークブックは KeyError → 500 で報告される。
    drive.configure_template()
    drive.template_bytes = make_template_bytes(sheet_name="別のシート")
    resp = client.post(REPORTS_URL, json={"rooms": []})

    assert resp.status_code == 500
    assert SHEET in resp.json()["error"]


def test_invalid_template_bytes_returns_500(client, drive):
    # xlsx でないバイト列は load_workbook が失敗し 500 で報告される。
    drive.configure_template()
    drive.template_bytes = b"not a real workbook"
    resp = client.post(REPORTS_URL, json={"rooms": []})

    assert resp.status_code == 500
    assert resp.json()["error"]


# --- 認証・CORS -------------------------------------------------------------

def test_request_without_token_returns_401(anon_client):
    resp = anon_client.post(REPORTS_URL, json={"rooms": []})

    assert resp.status_code == 401
    assert "サインイン" in resp.json()["error"]


def test_cors_preflight_for_local_dev(client):
    # ローカル開発（Vite dev サーバー）からのプリフライトに応答できる。
    resp = client.options(REPORTS_URL, headers={
        "Origin": "http://localhost:5173",
        "Access-Control-Request-Method": "POST",
        "Access-Control-Request-Headers": "Authorization, Content-Type",
    })

    assert resp.status_code == 200
    assert resp.headers["access-control-allow-origin"] == "http://localhost:5173"
