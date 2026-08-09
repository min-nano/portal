"""テスト用の共通ヘルパー。

雛形はリポジトリに同梱しない（社外秘フォーマット）ため、旧リポジトリの
テストと同様に最小限のワークブックをメモリ上で組み立てて使う。セル位置の
期待値は mapping.json から解決し、レイアウト変更に追従できるようにする。
"""

import io
import json
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection

MAPPING = json.load(
    open(
        os.path.join(os.path.dirname(__file__), "..", "app", "mapping.json"),
        encoding="utf-8",
    )
)
SHEET = MAPPING["sheet_name"]
BLOCK = MAPPING["room_block"]
STARTS = MAPPING["block_start_rows"]


def make_template_bytes(sheet_name=SHEET) -> bytes:
    """最小限の雛形ワークブックを組み立てて xlsx バイト列で返す。"""
    wb = Workbook()
    wb.active.title = sheet_name
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _unlock_block(ws, start_row):
    ws[f"{BLOCK['floor_col']}{start_row}"].protection = Protection(locked=False)
    ws[f"{BLOCK['room_name_col']}{start_row}"].protection = Protection(locked=False)
    for m in BLOCK["measurements"]:
        row = start_row + m["row_offset"]
        if "select" in m:
            ws[f"{m['select']['col']}{row}"].protection = Protection(locked=False)
        for col in BLOCK["value_fields"].values():
            ws[f"{col}{row}"].protection = Protection(locked=False)


def make_protected_template_bytes(unlocked_starts) -> bytes:
    """シート保護が有効で、指定ブロックのセルだけアンロックした雛形を作る。"""
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.protection.sheet = True
    ws[MAPPING["property_name_cell"]].protection = Protection(locked=False)
    for start_row in unlocked_starts:
        _unlock_block(ws, start_row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def decode_workbook(content: bytes):
    return load_workbook(io.BytesIO(content))


def measurement(measurement_key):
    return next(m for m in BLOCK["measurements"] if m["key"] == measurement_key)


def value_cell(room_index, measurement_key, field):
    """数値欄（diff/distance/digital_level）の書き込み先セルを mapping から解決する。"""
    start = STARTS[room_index]
    offset = measurement(measurement_key)["row_offset"]
    return f"{BLOCK['value_fields'][field]}{start + offset}"


def select_cell(room_index, measurement_key):
    """選択欄（傾斜方向 / 測定した壁・柱）の書き込み先セルを解決する。"""
    start = STARTS[room_index]
    m = measurement(measurement_key)
    return f"{m['select']['col']}{start + m['row_offset']}"
